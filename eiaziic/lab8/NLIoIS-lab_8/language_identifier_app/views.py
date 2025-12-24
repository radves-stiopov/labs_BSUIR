from django.shortcuts import render, get_object_or_404
from .utils.pdf_preprocessing import extract_text_from_pdf, clean_text
from .models import Document, AnalysisResult, LanguageProfile
import time
import io, csv
from django.http import HttpResponse
from .utils.lang_detection import (
    detect_language_freq,
    detect_language_short,
    detect_language_neural_ollama
)
import openpyxl
from openpyxl.styles import Font


def upload_files(request):
    if request.method == 'POST' and request.FILES.getlist('files'):
        uploaded_files = request.FILES.getlist('files')
        processed_files = []

        for f in uploaded_files:
            text = extract_text_from_pdf(f)
            if text:
                clean = clean_text(text)
                doc = Document.objects.create(name=f.name, text_content=clean)
                processed_files.append(doc)

        return render(request, 'language_app/result.html', {'files': processed_files})

    return render(request, 'language_app/index.html')


def result_view(request):
    files = Document.objects.order_by('-upload_date')[:10]
    return render(request, 'language_app/result.html', {'files': files})


def documents_list(request):
    documents = Document.objects.order_by('-upload_date')
    return render(request, 'language_app/documents_list.html', {'documents': documents})


def document_detail(request, doc_id):
    document = get_object_or_404(Document, id=doc_id)
    return render(request, 'language_app/document_detail.html', {'document': document})


def select_documents_for_analysis(request):
    documents = Document.objects.order_by('-upload_date')
    return render(request, 'language_app/select_documents.html', {'documents': documents})


def analyze_selected_documents(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('document_ids')
        documents = Document.objects.filter(id__in=selected_ids)
        profiles = LanguageProfile.objects.all()

        analysis = {}

        for doc in documents:
            text = doc.text_content
            results = []

            # --- Метод частотных слов ---
            start_time_freq = time.perf_counter()  # Засекаем время начала
            lang_freq, score_freq = detect_language_freq(text, profiles)
            processing_time_freq = time.perf_counter() - start_time_freq  # Считаем длительность

            r_freq = AnalysisResult.objects.create(
                document=doc,
                method='freq',
                detected_language=lang_freq,
                accuracy_score=score_freq,
                processing_time=processing_time_freq  # <-- Сохраняем время в БД
            )
            results.append(r_freq)

            # --- Метод коротких слов ---
            start_time_short = time.perf_counter()  # Засекаем время начала
            lang_short, score_short = detect_language_short(text, profiles)
            processing_time_short = time.perf_counter() - start_time_short  # Считаем длительность

            r_short = AnalysisResult.objects.create(
                document=doc,
                method='short',
                detected_language=lang_short,
                accuracy_score=score_short,
                processing_time=processing_time_short  # <-- Сохраняем время в БД
            )
            results.append(r_short)

            # --- Нейросетевой метод ---
            start_time_neural = time.perf_counter()  # Засекаем время начала
            lang_neural, score_neural = detect_language_neural_ollama(text, model_name='llama3')
            processing_time_neural = time.perf_counter() - start_time_neural  # Считаем длительность

            r_neural = AnalysisResult.objects.create(
                document=doc,
                method='neural',
                detected_language=lang_neural,
                accuracy_score=score_neural,
                processing_time=processing_time_neural  # <-- Сохраняем время в БД
            )
            results.append(r_neural)

            # Сохраняем результаты для шаблона
            analysis[doc] = results

        # Код для передачи ID для скачивания отчета остается без изменений
        doc_ids_query_string = '&'.join([f'ids={doc_id}' for doc_id in selected_ids])
        context = {
            'analysis': analysis,
            'doc_ids_query_string': doc_ids_query_string
        }

        return render(request, 'language_app/analysis_results.html', context)

    return render(request, 'language_app/select_documents.html', {'documents': Document.objects.all()})


def download_analysis_results(request):
    # Получаем список ID из GET-параметров URL
    selected_ids = request.GET.getlist('ids')

    if selected_ids:
        # Если ID переданы, фильтруем документы по этому списку
        documents = Document.objects.filter(id__in=selected_ids).order_by('upload_date')
    else:
        # Если кто-то открыл ссылку без ID, возвращаем пустой набор, чтобы не скачивать всё
        documents = Document.objects.none()


    # Создаем новую книгу Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Анализ документов"

    # Устанавливаем ширину столбцов
    column_width = 25
    worksheet.column_dimensions['A'].width = column_width * 2
    worksheet.column_dimensions['B'].width = column_width
    worksheet.column_dimensions['C'].width = column_width
    worksheet.column_dimensions['D'].width = column_width
    worksheet.column_dimensions['E'].width = column_width
    worksheet.column_dimensions['F'].width = column_width * 1.5

    # Заголовки
    headers = [
        'Документ', 'Определенный язык', 'Метрика (частотных слов)',
        'Метрика (коротких слов)', 'Метрика (нейросетевой)', 'Дата анализа'
    ]

    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Данные
    for row_idx, doc in enumerate(documents, 2):
        results = {r.method: r for r in doc.results.all()}

        main_language = 'N/A'
        if results.get('neural'):
            main_language = results.get('neural').detected_language
        elif results.get('freq'):
            main_language = results.get('freq').detected_language
        elif results.get('short'):
            main_language = results.get('short').detected_language

        row_data = [
            doc.name,
            main_language,
            f"{results.get('freq').accuracy_score:.4f}" if results.get('freq') and results.get(
                'freq').accuracy_score is not None else 'N/A',
            f"{results.get('short').accuracy_score:.4f}" if results.get('short') and results.get(
                'short').accuracy_score is not None else 'N/A',
            f"{results.get('neural').accuracy_score:.4f}" if results.get('neural') and results.get(
                'neural').accuracy_score is not None else 'N/A',
            doc.upload_date.strftime('%d.%m.%Y %H:%M')
        ]

        for col_idx, value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    # Сохраняем в буфер
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Создаем HTTP response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="analysis_results.xlsx"'
    response['Cache-Control'] = 'no-cache'

    return response


def analysis_history_view(request):
    documents_with_results = Document.objects.filter(
        results__isnull=False
    ).distinct().order_by('-upload_date')

    analysis = {}
    for doc in documents_with_results:
        analysis[doc] = doc.results.all().order_by('created_at')

    context = {
        'analysis': analysis
    }
    return render(request, 'language_app/analysis_history.html', context)


def help_view(request):
    return render(request, 'language_app/help.html')